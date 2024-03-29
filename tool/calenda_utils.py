#encoding:utf-8

import os,sys
if __name__ == '__main__':
    sys.path.insert(0,os.path.abspath(os.curdir))

#from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.style import Color
from openpyxl.style import Fill
from openpyxl.style import NumberFormat, Border, Color, Font
#from openpyxl.cell import Cell,get_column_letter
from openpyxl.cell import get_column_letter

import datetime

def make_calenda(task_list, date_from, days, save_as):

    date_to = date_from + datetime.timedelta(days=days)
    print date_to

    from_column = 2
    row = 0

    # Load the workbook...
    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.cell(row = 0, column = 0).value = u'任务名称'
    ws.cell(row = 0, column = 1).value = u'任务说明'

    month_set = set()
    weekday_column_list = []
    for i in range(days):
        d = date_from + datetime.timedelta(days=i)
        month = d.month
        if month not in month_set:
            cell = ws.cell(row = row, column = i+from_column)
            cell.value = u'%s月' % d.month
            month_set.add(month)
        cell = ws.cell(row = row+1, column = i+from_column)
        cell.value = '%s' % d.day        
        if d.weekday() in (5,6):
            weekday_column_list.append(i+from_column)

    for ws_column in range(from_column+1,from_column+days+1): 
        col_letter = get_column_letter(ws_column) 
        ###~ ws.column_dimensions[col_letter].auto_size = True 
        ws.column_dimensions[col_letter].width = 2.7

    for row in range(1,len(task_list)+2):
        for column in weekday_column_list:
            cell = ws.cell(row=row,column=column)
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = 'CBCACA'
            

    for i,task in enumerate(task_list):
        cell = ws.cell(row=i+2,column=0)
        cell.value = task

    #设置单元格border
    for row in range(0,len(task_list)+2):
        for col in range(0, days+2):
            cell = ws.cell(row=row,column=col)
            cell.style.borders.top.border_style = Border.BORDER_THIN
            cell.style.borders.right.border_style = Border.BORDER_THIN
            cell.style.borders.bottom.border_style = Border.BORDER_THIN
            cell.style.borders.left.border_style = Border.BORDER_THIN

    #设置第一列宽
    ws.column_dimensions['A'].width = 30

    wb.save(save_as)

if __name__ == '__main__':
    task_list1 = [
        u'需求分析',
        u'概要设计',
        u'详细设计',
        u'编码',
        u'内部测试',
        u'外部测试',
        u'部署',
        u'上线',
    ]

    task_days = 30
    date_from1 = datetime.datetime.now()
    save_as1 = 'calenda.xlsx'
    make_calenda(task_list1, date_from1, task_days, save_as1)
